#!/usr/bin/env python3
"""
Export Losing Trades to Excel
Exports all losing trades from a specific price band to an Excel file for analysis
"""

import pandas as pd
from pathlib import Path
import logging
import yaml
import sys
import os
import re
from typing import Optional

# Import Kite API utilities for CPR width calculation
# Calculate project root: backtesting/analytics/hybrid_exit -> backtesting/analytics -> backtesting -> root
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent.parent  # Go up 3 levels: hybrid_exit -> analytics -> backtesting -> root
ORIGINAL_CWD = os.getcwd()

# Add project root to Python path
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

try:
    from trading_bot_utils import get_kite_api_instance
except (ImportError, FileNotFoundError) as e:
    # Try changing to project root directory
    os.chdir(PROJECT_ROOT)
    try:
        from trading_bot_utils import get_kite_api_instance
    finally:
        os.chdir(ORIGINAL_CWD)

try:
    from calculate_high_swing_low import (
        calculate_high_between_entry_exit,
        calculate_swing_low_at_entry,
    )
except ImportError:
    # Ensure analytics directory (parent of hybrid_exit) is on sys.path
    # calculate_high_swing_low.py is in backtesting/analytics/, not backtesting/analytics/hybrid_exit/
    analytics_dir = Path(__file__).parent.parent  # Go up one level from hybrid_exit to analytics
    if str(analytics_dir) not in sys.path:
        sys.path.insert(0, str(analytics_dir))
    from calculate_high_swing_low import (
        calculate_high_between_entry_exit,
        calculate_swing_low_at_entry,
    )

_cached_kite_client = None
_skip_first_flag_cache = {}
_strategy_df_cache = {}

SENTIMENT_COLUMNS = [
    'nifty_prev_day_close',
    'nifty_open_price',
    'nifty_930',
    'nifty_current_price',
    'nifty_sentiment',
    'nifty_open_minus_current',
    'nifty_open_sentiment',
    'nifty_930_minus_current',
    'current_price_vs_pivot',
]

ALWAYS_ON_SENTIMENT_COLUMNS = [
    'nifty_930_sentiment',
    'pivot_sentiment'
]

INDICATOR_COLUMNS = [
    'supertrend1',
    'supertrend1_dir',
    'k',
    'd',
    'fast_wpr',
    'slow_wpr',
    'fast_ma',
    'slow_ma'
]

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Price band to filter (70-249)
PRICE_BAND_MIN = 70
PRICE_BAND_MAX = 249
# PnL threshold: trades below this percentage are considered losing
PNL_THRESHOLD_LOSS = -2.5

def load_config(config_path: Path) -> dict:
    """Load YAML config safely."""
    if not config_path or not config_path.exists():
        return {}
    try:
        with open(config_path, 'r') as f:
            return yaml.safe_load(f) or {}
    except Exception as e:
        logger.warning(f"Unable to load config {config_path}: {e}")
        return {}

def analytics_sentiment_enabled(config: dict) -> bool:
    """Return True if sentiment columns should be included in analytics output."""
    return config.get('ANALYTICS_OUTPUT', {}).get('INCLUDE_SENTIMENT_COLUMNS', True)

def get_swing_low_candles(config: dict) -> int:
    """Read swing low candle setting from config."""
    return config.get('FIXED', {}).get('SWING_LOW_CANDLES', 5)

def extract_date_from_file_path(file_path: Path) -> Optional[str]:
    """Extract date from trade file path and convert to YYYY-MM-DD format.
    
    File path format: {expiry_week}_DYNAMIC/{day_label}/entry2_dynamic_atm_mkt_sentiment_trades.csv
    Example: NOV18_DYNAMIC/NOV26/entry2_dynamic_atm_mkt_sentiment_trades.csv -> 2025-11-26
    """
    try:
        # Get the parent directory name (day_label like NOV26)
        day_label = file_path.parent.name
        
        # Parse day_label (e.g., NOV26 -> month=NOV, day=26)
        if len(day_label) >= 5:
            month_str = day_label[:3].upper()  # NOV
            day_str = day_label[3:].lstrip('0')  # 26 (remove leading zeros)
            
            # Month mapping
            month_map = {
                'JAN': '01', 'FEB': '02', 'MAR': '03', 'APR': '04',
                'MAY': '05', 'JUN': '06', 'JUL': '07', 'AUG': '08',
                'SEP': '09', 'OCT': '10', 'NOV': '11', 'DEC': '12'
            }
            
            month_num = month_map.get(month_str)
            if month_num and day_str:
                # Assume year 2025 (can be made configurable if needed)
                date_str = f"2025-{month_num}-{day_str.zfill(2)}"
                # Validate the date
                pd.to_datetime(date_str)
                return date_str
    except Exception as e:
        logger.debug(f"Could not extract date from path {file_path}: {e}")
    
    return None

def normalize_time_str(time_value: Optional[str]) -> Optional[str]:
    """Ensure time string includes seconds (HH:MM:SS)."""
    if not time_value or not isinstance(time_value, str):
        return None
    time_value = time_value.strip()
    if not time_value:
        return None
    if len(time_value) == 5:
        return f"{time_value}:00"
    return time_value

def extract_hyperlink_path(value: str) -> Optional[str]:
    """Pull path from Excel-style HYPERLINK string."""
    if not isinstance(value, str):
        return None
    match = re.search(r'=HYPERLINK\("([^"]+)"', value.replace('""', '"'))
    if match:
        return match.group(1)
    return None

def resolve_strategy_file(symbol_value, symbol_html_value, source_file: Optional[str]) -> Optional[Path]:
    """Best-effort resolution of strategy CSV path for a trade row."""
    # 1. Direct hyperlink on symbol column
    hyperlink_path = extract_hyperlink_path(symbol_value) if symbol_value else None
    if hyperlink_path:
        return Path(hyperlink_path)
    
    # 2. Hyperlink on symbol_html column (HTML file) -> convert to CSV
    html_path = extract_hyperlink_path(symbol_html_value) if symbol_html_value else None
    if html_path:
        csv_candidate = Path(html_path).with_suffix('.csv')
        return csv_candidate
    
    # 3. Construct from symbol text and source directory
    symbol_text = None
    if isinstance(symbol_value, str):
        if '=HYPERLINK' in symbol_value:
            text_match = re.findall(r'"([^"]+)"', symbol_value.replace('""', '"'))
            if len(text_match) >= 2:
                symbol_text = text_match[1]
        else:
            symbol_text = symbol_value.strip()
    
    if not symbol_text:
        return None
    
    strategy_filename = symbol_text
    if not strategy_filename.endswith('_strategy.csv'):
        strategy_filename = f"{strategy_filename}_strategy.csv"
    
    if source_file:
        trade_dir = Path(source_file).parent
        atm_dir = trade_dir / 'ATM'
        candidate_dirs = [atm_dir, trade_dir]
    else:
        candidate_dirs = []
    
    for directory in candidate_dirs:
        candidate_path = directory / strategy_filename
        if candidate_path.exists():
            return candidate_path
    
    # Fall back to direct filename if reachable
    candidate_path = Path(strategy_filename)
    return candidate_path if candidate_path.exists() else None

def _build_skip_first_mapping(strategy_file: Path) -> dict:
    """Create a mapping of entry_time (HH:MM:SS) -> skip_first flag (0/1) for a strategy file."""
    mapping = {}
    
    try:
        df = pd.read_csv(strategy_file)
    except Exception as e:
        logger.debug(f"Could not read strategy file {strategy_file} for skip_first mapping: {e}")
        return mapping
    
    if df.empty or 'entry2_entry_type' not in df.columns or 'supertrend1_dir' not in df.columns:
        return mapping
    
    if 'date' not in df.columns:
        logger.debug(f"'date' column missing in {strategy_file}, cannot compute skip_first mapping")
        return mapping
    
    try:
        df['date'] = pd.to_datetime(df['date'])
    except Exception:
        logger.debug(f"Could not parse dates in {strategy_file}, cannot compute skip_first mapping")
        return mapping
    
    prev_dirs = df['supertrend1_dir'].shift(1)
    skip_flag = False
    
    for idx, row in df.iterrows():
        prev_dir = prev_dirs.iloc[idx]
        curr_dir = row.get('supertrend1_dir')
        
        if prev_dir == 1 and curr_dir == -1:
            skip_flag = True
        
        entry_type = row.get('entry2_entry_type')
        if isinstance(entry_type, str) and entry_type.strip().upper() == 'ENTRY':
            entry_time = row.get('date')
            if pd.notna(entry_time):
                time_key = entry_time.strftime('%H:%M:%S')
                mapping[time_key] = 1 if skip_flag else 0
            skip_flag = False
    
    return mapping

def get_skip_first_mapping(strategy_file: Path) -> dict:
    """Retrieve cached skip_first mapping for a strategy file."""
    cache_key = str(strategy_file)
    if cache_key in _skip_first_flag_cache:
        return _skip_first_flag_cache[cache_key]
    
    if not strategy_file.exists():
        _skip_first_flag_cache[cache_key] = {}
        return _skip_first_flag_cache[cache_key]
    
    mapping = _build_skip_first_mapping(strategy_file)
    _skip_first_flag_cache[cache_key] = mapping
    return mapping

def get_nifty_file_path_from_trade_file(trade_file_path: Optional[str], date_str: Optional[str]) -> Optional[Path]:
    """Get NIFTY50 1min data file path from trade file path or date."""
    if trade_file_path:
        # Extract from trade file path: {expiry}_DYNAMIC/{day_label}/entry2_dynamic_atm_mkt_sentiment_trades.csv
        # NIFTY file: {expiry}_DYNAMIC/{day_label}/nifty50_1min_data_{day_label_lower}.csv
        trade_path = Path(trade_file_path)
        day_label = trade_path.parent.name  # e.g., NOV26
        day_label_lower = day_label.lower()  # nov26
        expiry_dir = trade_path.parent.parent  # {expiry}_DYNAMIC
        nifty_file = expiry_dir / day_label / f"nifty50_1min_data_{day_label_lower}.csv"
        if nifty_file.exists():
            return nifty_file
    
    # Try to construct from date if available
    if date_str:
        try:
            date_obj = pd.to_datetime(date_str)
            day_label = date_obj.strftime('%b%d').upper()  # NOV26
            day_label_lower = day_label.lower()  # nov26
            
            # Try to find in data directory
            script_dir = Path(__file__).parent
            possible_data_paths = [
                script_dir.parent / 'data',
                script_dir.parent.parent / 'backtesting' / 'data',
                Path('data'),
                Path('backtesting/data'),
            ]
            
            for data_dir in possible_data_paths:
                if data_dir.exists():
                    # Search in all expiry directories
                    for expiry_dir in data_dir.glob('*_DYNAMIC'):
                        nifty_file = expiry_dir / day_label / f"nifty50_1min_data_{day_label_lower}.csv"
                        if nifty_file.exists():
                            return nifty_file
        except Exception as e:
            logger.debug(f"Could not construct NIFTY file path from date {date_str}: {e}")
    
    return None

def get_nifty_price_at_time(nifty_file: Path, time_str: str) -> Optional[float]:
    """Get NIFTY50 close price at a specific time from 1min data file."""
    if not nifty_file or not nifty_file.exists():
        return None
    
    try:
        time_normalized = normalize_time_str(time_str)
        if not time_normalized:
            return None
        
        df = pd.read_csv(nifty_file)
        if df.empty or 'date' not in df.columns or 'close' not in df.columns:
            return None
        
        df['date'] = pd.to_datetime(df['date'])
        df['time'] = df['date'].dt.time
        
        # Parse the time string (HH:MM:SS)
        time_parts = time_normalized.split(':')
        if len(time_parts) >= 2:
            target_hour = int(time_parts[0])
            target_minute = int(time_parts[1])
            target_time = pd.Timestamp.now().replace(hour=target_hour, minute=target_minute, second=0, microsecond=0).time()
            
            # Find the closest time (exact match or next available)
            matching_rows = df[df['time'] == target_time]
            if matching_rows.empty:
                # Find the next available time after target
                later_rows = df[df['time'] > target_time]
                if not later_rows.empty:
                    matching_rows = later_rows.head(1)
                else:
                    # Find the closest earlier time
                    earlier_rows = df[df['time'] < target_time].tail(1)
                    if not earlier_rows.empty:
                        matching_rows = earlier_rows
            
            if not matching_rows.empty:
                close_price = matching_rows.iloc[0]['close']
                if pd.notna(close_price):
                    return float(close_price)
    except Exception as e:
        logger.debug(f"Error getting NIFTY price at {time_str} from {nifty_file}: {e}")
    
    return None

def get_nifty_open_price(nifty_file: Path) -> Optional[float]:
    """Get NIFTY50 open price (first candle of the trading day, typically 9:15 or 9:16)."""
    if not nifty_file or not nifty_file.exists():
        return None
    
    try:
        df = pd.read_csv(nifty_file)
        if df.empty or 'date' not in df.columns or 'open' not in df.columns:
            return None
        
        df['date'] = pd.to_datetime(df['date'])
        df['time'] = df['date'].dt.time
        
        # Get the first row (should be 9:15 or 9:16)
        first_row = df.iloc[0]
        open_price = first_row.get('open')
        if pd.notna(open_price):
            return float(open_price)
    except Exception as e:
        logger.debug(f"Error getting NIFTY open price from {nifty_file}: {e}")
    
    return None

def get_nifty_price_at_930(nifty_file: Path) -> Optional[float]:
    """Get NIFTY50 close price at 9:30am."""
    return get_nifty_price_at_time(nifty_file, "09:30:00")

def get_nifty_prev_day_close(nifty_file: Path) -> Optional[float]:
    """Get previous trading day's close price for NIFTY50."""
    if not nifty_file or not nifty_file.exists():
        return None
    
    try:
        # Try using Kite API first (more reliable)
        prev_day_high, prev_day_low, prev_day_close, prev_day_date = fetch_prev_day_nifty_ohlc_via_kite(str(nifty_file))
        if prev_day_close is not None:
            return prev_day_close
        
        # Fallback: try to read from file (if previous day data is in the file)
        df = pd.read_csv(nifty_file)
        if df.empty or 'date' not in df.columns:
            return None
        
        df['date'] = pd.to_datetime(df['date'])
        # Get the first date in the file
        first_date = df['date'].iloc[0].date()
        
        # The file might contain previous day's data at the beginning
        # Look for data before the first date
        from datetime import timedelta
        prev_date = first_date - timedelta(days=1)
        
        # Check if there's data for previous date
        prev_date_str = prev_date.strftime('%Y-%m-%d')
        prev_rows = df[df['date'].dt.date == prev_date]
        if not prev_rows.empty and 'close' in prev_rows.columns:
            # Get the last close of previous day
            last_close = prev_rows.iloc[-1]['close']
            if pd.notna(last_close):
                return float(last_close)
    except Exception as e:
        logger.debug(f"Error getting previous day close from {nifty_file}: {e}")
    
    return None

def get_nifty_prev_day_pivot(nifty_file: Path) -> Optional[float]:
    """Calculate previous day's pivot ((H + L + C) / 3)."""
    if not nifty_file or not nifty_file.exists():
        return None
    
    try:
        prev_high, prev_low, prev_close, _ = fetch_prev_day_nifty_ohlc_via_kite(str(nifty_file))
        if prev_high is not None and prev_low is not None and prev_close is not None:
            return (prev_high + prev_low + prev_close) / 3
    except Exception as e:
        logger.debug(f"Error fetching prev day OHLC for pivot via Kite: {e}")
    
    try:
        df = pd.read_csv(nifty_file)
        if df.empty or 'date' not in df.columns:
            return None
        df['date'] = pd.to_datetime(df['date'])
        first_date = df['date'].iloc[0].date()
        from datetime import timedelta
        prev_date = first_date - timedelta(days=1)
        prev_rows = df[df['date'].dt.date == prev_date]
        if prev_rows.empty:
            return None
        required_cols = {'high', 'low', 'close'}
        if not required_cols.issubset(prev_rows.columns):
            return None
        prev_high = prev_rows['high'].max()
        prev_low = prev_rows['low'].min()
        prev_close = prev_rows.iloc[-1]['close']
        if pd.notna(prev_high) and pd.notna(prev_low) and pd.notna(prev_close):
            return (float(prev_high) + float(prev_low) + float(prev_close)) / 3
    except Exception as e:
        logger.debug(f"Error calculating pivot from file {nifty_file}: {e}")
    
    return None

def enrich_with_nifty_data(df: pd.DataFrame, enable_sentiment: bool = True) -> pd.DataFrame:
    """Add nifty_prev_day_close, nifty_current_price, nifty_sentiment, nifty_open_price, 
    nifty_930, and difference columns."""
    if df.empty:
        return df
    
    nifty_prev_close_values = []
    nifty_current_price_values = []
    nifty_sentiment_values = []
    nifty_open_price_values = []
    nifty_930_values = []
    nifty_open_minus_current_values = []
    nifty_930_minus_current_values = []
    nifty_open_sentiment_values = []
    nifty_930_sentiment_values = []
    current_vs_pivot_values = []
    pivot_sentiment_values = []
    
    for idx, row in df.iterrows():
        trade_file_path = row.get('source_trade_file')
        date_str = row.get('date')
        entry_time = normalize_time_str(row.get('entry_time'))
        
        nifty_file = get_nifty_file_path_from_trade_file(trade_file_path, date_str)
        
        nifty_prev_close = None
        nifty_current_price = None
        nifty_sentiment = None
        nifty_open_price = None
        nifty_930 = None
        nifty_open_minus_current = None
        nifty_930_minus_current = None
        nifty_open_sentiment = None
        nifty_930_sentiment = None
        pivot_value = None
        current_vs_pivot = None
        pivot_sentiment = None
        
        if nifty_file and nifty_file.exists():
            # Get previous day close
            nifty_prev_close = get_nifty_prev_day_close(nifty_file)
            pivot_value = get_nifty_prev_day_pivot(nifty_file)
            
            # Get open price (first candle of the day)
            nifty_open_price = get_nifty_open_price(nifty_file)
            
            # Get price at 9:30am
            nifty_930 = get_nifty_price_at_930(nifty_file)
            
            # Get current price at entry time
            if entry_time:
                nifty_current_price = get_nifty_price_at_time(nifty_file, entry_time)
            
            # Calculate sentiment: prev_close - current_price
            # Positive = price went down (bearish), Negative = price went up (bullish)
            if nifty_prev_close is not None and nifty_current_price is not None:
                diff = nifty_prev_close - nifty_current_price
                if diff > 0:
                    nifty_sentiment = "BEARISH"  # Price went down
                elif diff < 0:
                    nifty_sentiment = "BULLISH"  # Price went up
                else:
                    nifty_sentiment = "NEUTRAL"
            
            # Calculate differences and sentiments
            if nifty_open_price is not None and nifty_current_price is not None:
                nifty_open_minus_current = nifty_open_price - nifty_current_price
                # If current > open (diff is negative), price went UP → BULLISH
                # If current < open (diff is positive), price went DOWN → BEARISH
                if nifty_open_minus_current < 0:
                    nifty_open_sentiment = "BULLISH"  # Price went up from open
                elif nifty_open_minus_current > 0:
                    nifty_open_sentiment = "BEARISH"  # Price went down from open
                else:
                    nifty_open_sentiment = "NEUTRAL"
            
            if nifty_930 is not None and nifty_current_price is not None:
                nifty_930_minus_current = nifty_930 - nifty_current_price
                # If current > 9:30 (diff is negative), price went UP → BULLISH
                # If current < 9:30 (diff is positive), price went DOWN → BEARISH
                if nifty_930_minus_current < 0:
                    nifty_930_sentiment = "BULLISH"  # Price went up from 9:30
                elif nifty_930_minus_current > 0:
                    nifty_930_sentiment = "BEARISH"  # Price went down from 9:30
                else:
                    nifty_930_sentiment = "NEUTRAL"
            
            if pivot_value is not None and nifty_current_price is not None:
                current_vs_pivot = nifty_current_price - pivot_value
                if current_vs_pivot > 0:
                    pivot_sentiment = "BULLISH"
                elif current_vs_pivot < 0:
                    pivot_sentiment = "BEARISH"
                else:
                    pivot_sentiment = "NEUTRAL"
        
        nifty_prev_close_values.append(nifty_prev_close)
        nifty_current_price_values.append(nifty_current_price)
        nifty_sentiment_values.append(nifty_sentiment)
        nifty_open_price_values.append(nifty_open_price)
        nifty_930_values.append(nifty_930)
        nifty_open_minus_current_values.append(nifty_open_minus_current)
        nifty_930_minus_current_values.append(nifty_930_minus_current)
        nifty_open_sentiment_values.append(nifty_open_sentiment)
        nifty_930_sentiment_values.append(nifty_930_sentiment)
        current_vs_pivot_values.append(current_vs_pivot)
        pivot_sentiment_values.append(pivot_sentiment)
    
    if enable_sentiment:
        df['nifty_prev_day_close'] = nifty_prev_close_values
        df['nifty_current_price'] = nifty_current_price_values
        df['nifty_sentiment'] = nifty_sentiment_values
        df['nifty_open_price'] = nifty_open_price_values
        df['nifty_930'] = nifty_930_values
        df['nifty_open_minus_current'] = nifty_open_minus_current_values
        df['nifty_930_minus_current'] = nifty_930_minus_current_values
        df['nifty_open_sentiment'] = nifty_open_sentiment_values
        df['current_price_vs_pivot'] = current_vs_pivot_values
    
    df['nifty_930_sentiment'] = nifty_930_sentiment_values
    df['pivot_sentiment'] = pivot_sentiment_values
    
    if not enable_sentiment:
        drop_cols = [col for col in SENTIMENT_COLUMNS if col in df.columns]
        if drop_cols:
            df = df.drop(columns=drop_cols)
    
    return df

def enrich_with_high_swing_low(df: pd.DataFrame, config: dict) -> pd.DataFrame:
    """Add/compute high and swing_low columns using strategy data."""
    if df.empty:
        return df

    swing_low_candles = get_swing_low_candles(config)
    high_pct_values = []
    swing_low_pct_values = []
    high_abs_values = []
    swing_low_abs_values = []
    
    for idx, row in df.iterrows():
        existing_high = row.get('high')
        existing_swing_low = row.get('swing_low')
        
        # Preserve pre-computed values when present
        if pd.notna(existing_high) and pd.notna(existing_swing_low):
            high_abs_values.append(existing_high)
            swing_low_abs_values.append(existing_swing_low)
            # Calculate percentages
            entry_price_val = row.get('entry_price')
            if pd.notna(entry_price_val) and entry_price_val:
                try:
                    entry_price_float = float(entry_price_val)
                    high_pct = ((existing_high - entry_price_float) / entry_price_float * 100) if existing_high is not None else None
                    swing_low_pct = ((existing_swing_low - entry_price_float) / entry_price_float * 100) if existing_swing_low is not None else None
                except Exception:
                    high_pct = None
                    swing_low_pct = None
            else:
                high_pct = None
                swing_low_pct = None
            high_pct_values.append(high_pct)
            swing_low_pct_values.append(swing_low_pct)
            continue
        
        entry_time = normalize_time_str(row.get('entry_time'))
        exit_time = normalize_time_str(row.get('exit_time'))
        entry_price = row.get('entry_price')
        strategy_file = resolve_strategy_file(
            row.get('symbol'),
            row.get('symbol_html'),
            row.get('source_trade_file')
        )
        
        high = existing_high if pd.notna(existing_high) else None
        swing_low = existing_swing_low if pd.notna(existing_swing_low) else None
        
        if strategy_file and strategy_file.exists() and entry_time and exit_time and pd.notna(entry_price):
            try:
                entry_price_float = float(entry_price)
                if high is None:
                    high = calculate_high_between_entry_exit(strategy_file, entry_time, exit_time, entry_price_float)
                if swing_low is None:
                    swing_low = calculate_swing_low_at_entry(strategy_file, entry_time, entry_price_float, swing_low_candles)
            except Exception as calc_err:
                logger.warning(f"High/Swing low calc failed for row {idx} ({strategy_file}): {calc_err}")
        else:
            if not strategy_file or not strategy_file.exists():
                logger.debug(f"Strategy file missing for row {idx}: {strategy_file}")
        
        high_abs_values.append(high)
        swing_low_abs_values.append(swing_low)

        entry_price_val = row.get('entry_price')
        if pd.notna(entry_price_val) and entry_price_val:
            try:
                entry_price_float = float(entry_price_val)
                high_pct = ((high - entry_price_float) / entry_price_float * 100) if high is not None else None
                swing_low_pct = ((swing_low - entry_price_float) / entry_price_float * 100) if swing_low is not None else None
            except Exception:
                high_pct = None
                swing_low_pct = None
        else:
            high_pct = None
            swing_low_pct = None
        high_pct_values.append(high_pct)
        swing_low_pct_values.append(swing_low_pct)

    # Replace high/swing_low columns with percentage change, but keep absolute values for reference
    df['high_abs'] = high_abs_values
    df['swing_low_abs'] = swing_low_abs_values
    df['high'] = high_pct_values
    df['swing_low'] = swing_low_pct_values
    return df

def add_skip_first_flags(df: pd.DataFrame) -> pd.DataFrame:
    """Add skip_first column (1 if entry is first after ST switch, else 0)."""
    if df.empty:
        df['skip_first'] = None
        return df
    
    skip_values = []
    
    for idx, row in df.iterrows():
        entry_time = normalize_time_str(row.get('entry_time'))
        strategy_file = resolve_strategy_file(
            row.get('symbol'),
            row.get('symbol_html'),
            row.get('source_trade_file')
        )
        
        skip_flag_value = None
        if entry_time and strategy_file and strategy_file.exists():
            mapping = get_skip_first_mapping(strategy_file)
            if mapping:
                skip_flag_value = mapping.get(entry_time, 0)
            else:
                skip_flag_value = 0
        
        skip_values.append(skip_flag_value)
    
    df['skip_first'] = skip_values
    return df

def _get_strategy_dataframe(strategy_file: Path) -> Optional[pd.DataFrame]:
    """Load strategy dataframe with caching."""
    cache_key = str(strategy_file)
    if cache_key in _strategy_df_cache:
        return _strategy_df_cache[cache_key]
    
    try:
        df = pd.read_csv(strategy_file)
        if df.empty or 'date' not in df.columns:
            _strategy_df_cache[cache_key] = None
            return None
        df['date'] = pd.to_datetime(df['date'])
        df = df.reset_index(drop=True)
        _strategy_df_cache[cache_key] = df
        return df
    except Exception as e:
        logger.debug(f"Could not read strategy file {strategy_file} for indicators: {e}")
        _strategy_df_cache[cache_key] = None
        return None

def _extract_indicator_values(strategy_df: pd.DataFrame, entry_time: str) -> dict:
    """Fetch indicator columns for the row matching entry_time."""
    result = {col: None for col in INDICATOR_COLUMNS}
    if strategy_df is None or entry_time is None:
        return result
    
    normalized_time = normalize_time_str(entry_time)
    if not normalized_time:
        return result
    
    try:
        target_time = pd.to_datetime(normalized_time).time()
    except Exception:
        return result
    
    # Strategy DF is single day, match on time
    matches = strategy_df[strategy_df['date'].dt.time == target_time]
    if matches.empty:
        return result
    
    idx = matches.index[0]
    row_idx = idx - 1 if idx > 0 else idx
    row = strategy_df.iloc[row_idx]
    for col in INDICATOR_COLUMNS:
        if col in row:
            result[col] = row[col]
    return result

def add_indicator_snapshot(df: pd.DataFrame) -> pd.DataFrame:
    """Append indicator columns at entry time using strategy files."""
    if df.empty:
        for col in INDICATOR_COLUMNS:
            df[col] = None
        return df
    
    indicator_data = {col: [] for col in INDICATOR_COLUMNS}
    
    for _, row in df.iterrows():
        entry_time = normalize_time_str(row.get('entry_time'))
        strategy_file = resolve_strategy_file(
            row.get('symbol'),
            row.get('symbol_html'),
            row.get('source_trade_file')
        )
        
        indicator_values = {col: None for col in INDICATOR_COLUMNS}
        if entry_time and strategy_file and strategy_file.exists():
            strat_df = _get_strategy_dataframe(strategy_file)
            indicator_values = _extract_indicator_values(strat_df, entry_time)
        
        for col in INDICATOR_COLUMNS:
            indicator_data[col].append(indicator_values[col])
    
    for col in INDICATOR_COLUMNS:
        df[col] = indicator_data[col]
    
    return df

def fetch_prev_day_nifty_ohlc_via_kite(csv_file_path: str):
    """Fetch previous trading day's OHLC data for NIFTY 50 using KiteConnect API"""
    global _cached_kite_client
    
    try:
        if _cached_kite_client is None:
            original_cwd = os.getcwd()
            try:
                os.chdir(PROJECT_ROOT)
                kite, _, _ = get_kite_api_instance(suppress_logs=True)
            finally:
                os.chdir(original_cwd)
            _cached_kite_client = kite
        else:
            kite = _cached_kite_client
        
        if kite is None:
            logger.debug("Could not get Kite API instance - cannot fetch previous day OHLC")
            return None, None, None, None
        
        df_tmp = pd.read_csv(csv_file_path)
        df_tmp['date'] = pd.to_datetime(df_tmp['date'])
        current_date = df_tmp['date'].iloc[0].date()
        from datetime import timedelta
        prev_date = current_date - timedelta(days=1)
        
        backoff_date = prev_date
        data = []
        for days_back in range(7):
            try:
                data = kite.historical_data(
                    instrument_token=256265,
                    from_date=backoff_date,
                    to_date=backoff_date,
                    interval='day'
                )
                if data and len(data) > 0:
                    logger.debug(f"Found trading day data for {backoff_date} (checked {days_back + 1} days back)")
                    break
            except Exception as e:
                logger.debug(f"Error fetching data for {backoff_date}: {e}")
            
            backoff_date = backoff_date - timedelta(days=1)
        
        if not data or len(data) == 0:
            logger.debug(f"Could not fetch previous day OHLC data for {current_date}")
            return None, None, None, None
        
        c = data[0]
        return float(c['high']), float(c['low']), float(c['close']), backoff_date
    except Exception as e:
        logger.debug(f"Error in fetch_prev_day_nifty_ohlc_via_kite: {e}")
        import traceback
        logger.debug(traceback.format_exc())
        return None, None, None, None

def calculate_cpr_width(data_dir: Path) -> float:
    """Calculate CPR width = TC - BC"""
    # Extract day_label from data_dir (e.g., NOV06 from NOV11_DYNAMIC/NOV06)
    day_label = data_dir.name.upper()
    day_label_lower = day_label.lower()
    nifty_file = data_dir / f"nifty50_1min_data_{day_label_lower}.csv"
    
    if not nifty_file.exists():
        logger.debug(f"Could not find nifty50_1min_data_{day_label_lower}.csv in {data_dir} - cannot calculate CPR width")
        return None
    
    try:
        prev_day_high, prev_day_low, prev_day_close, prev_day_date = fetch_prev_day_nifty_ohlc_via_kite(str(nifty_file))
        
        if prev_day_high is None or prev_day_low is None or prev_day_close is None:
            logger.debug(f"Could not fetch previous day OHLC for {data_dir.name} - cannot calculate CPR width")
            return None
        
        pivot = (prev_day_high + prev_day_low + prev_day_close) / 3
        bc = (prev_day_high + prev_day_low) / 2
        tc = (pivot - bc) + pivot
        cpr_width = abs(tc - bc)
        
        logger.debug(f"CPR width for {data_dir.name}: {cpr_width:.2f}")
        return cpr_width
        
    except Exception as e:
        logger.warning(f"Error calculating CPR width for {data_dir.name}: {e}")
        return None

def get_trade_files(config_file: Path):
    """Get all dynamic ATM trade files, excluding dates with CPR width > 60"""
    if not config_file.exists():
        logger.error(f"Config file not found: {config_file}")
        return []
    
    with open(config_file, 'r') as f:
        config = yaml.safe_load(f)
    
    expiry_config = config.get('BACKTESTING_EXPIRY', {})
    expiry_weeks = expiry_config.get('EXPIRY_WEEK_LABELS', [])
    backtesting_days = expiry_config.get('BACKTESTING_DAYS', [])
    
    def date_to_day_label(date_str):
        try:
            date_obj = pd.to_datetime(date_str)
            month = date_obj.strftime('%b').upper()
            day = date_obj.strftime('%d')
            if int(day) > 9:
                day = day.lstrip('0')
            return f"{month}{day}"
        except:
            return None
    
    trade_files = []
    filtered_days = []
    
    # Determine data directory base path
    # Script is at: backtesting/analytics/hybrid_exit/export_losing_trades.py
    # Data is at: backtesting/data/
    script_dir = Path(__file__).parent if '__file__' in globals() else Path.cwd()
    possible_data_paths = [
        script_dir.parent.parent / 'data',  # Go up 2 levels: hybrid_exit -> analytics -> backtesting, then /data
        script_dir.parent.parent.parent / 'backtesting' / 'data',  # From project root: backtesting/data
        Path('data'),  # Current directory
        Path('backtesting/data'),  # backtesting/ subdirectory
        PROJECT_ROOT / 'backtesting' / 'data',  # Use PROJECT_ROOT we calculated earlier
    ]
    
    data_dir_base = None
    for path in possible_data_paths:
        if path.exists():
            data_dir_base = path
            logger.debug(f"Found data directory at: {data_dir_base}")
            break
    
    if data_dir_base is None:
        logger.error(f"Data directory not found. Tried: {possible_data_paths}")
        return []
    
    for expiry_week in expiry_weeks:
        for date_str in backtesting_days:
            day_label = date_to_day_label(date_str)
            if not day_label:
                continue
            
            dynamic_path = data_dir_base / f"{expiry_week}_DYNAMIC" / day_label
            
            if dynamic_path.exists():
                cpr_width = calculate_cpr_width(dynamic_path)
                if cpr_width is not None and cpr_width > 60:
                    logger.info(f"[FILTER] FILTERING OUT {day_label} - CPR width ({cpr_width:.2f}) > 60")
                    filtered_days.append(f"{expiry_week}/{day_label}")
                    continue
                elif cpr_width is None:
                    logger.warning(f"[FILTER] Could not calculate CPR width for {day_label} - EXCLUDING")
                    filtered_days.append(f"{expiry_week}/{day_label}")
                    continue
                else:
                    logger.debug(f"[INCLUDE] Including {day_label} - CPR width ({cpr_width:.2f}) <= 60")
            
            dynamic_base = data_dir_base / f"{expiry_week}_DYNAMIC" / day_label
            trade_files.append(dynamic_base / 'entry2_dynamic_atm_mkt_sentiment_trades.csv')
    
    if filtered_days:
        logger.info(f"\n[FILTER SUMMARY] Filtered out {len(filtered_days)} days (CPR width > 60): {filtered_days}")
    else:
        logger.info(f"\n[OK] All days included (CPR width <= 60)")
    
    return trade_files

def main():
    # Determine config file path - try multiple locations
    script_dir = Path(__file__).parent
    possible_config_paths = [
        script_dir.parent / 'backtesting_config.yaml',  # backtesting/backtesting_config.yaml
        script_dir.parent.parent / 'backtesting_config.yaml',  # backtesting_config.yaml (root)
        Path('backtesting_config.yaml'),  # Current directory
        Path('backtesting/backtesting_config.yaml'),  # backtesting/ subdirectory
    ]
    
    config_file = None
    for path in possible_config_paths:
        if path.exists():
            config_file = path
            logger.debug(f"Found config file at: {config_file}")
            break
    
    if config_file is None:
        logger.error(f"Config file not found. Tried: {possible_config_paths}")
        return
    
    # Output directory (relative to script location)
    output_dir = script_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Load config once for downstream helpers
    config_data = load_config(config_file)
    
    # Get all trade files
    logger.info("Loading trade files from backtesting_config.yaml...")
    trade_files = get_trade_files(config_file)
    logger.info(f"Found {len(trade_files)} potential trade files")
    
    # Load and filter trades
    all_trades = []
    
    for trade_file in trade_files:
        if not trade_file.exists():
            logger.debug(f"File not found (skipping): {trade_file}")
            continue
        
        try:
            df = pd.read_csv(trade_file)
            if df.empty:
                logger.debug(f"File is empty (skipping): {trade_file}")
                continue
            
            # Ensure required columns exist
            required_cols = ['entry_price', 'pnl']
            if not all(col in df.columns for col in required_cols):
                logger.warning(f"Missing required columns in {trade_file.name}, skipping")
                continue
            
            # Filter out rows with missing entry_price or pnl
            df = df.dropna(subset=['entry_price', 'pnl'])
            
            # Convert to numeric
            df['entry_price'] = pd.to_numeric(df['entry_price'], errors='coerce')
            df['pnl'] = pd.to_numeric(df['pnl'], errors='coerce')
            
            # Remove rows with invalid values
            df = df[df['entry_price'].notna() & df['pnl'].notna()]
            
            # Filter for price band 70-249
            df = df[(df['entry_price'] >= PRICE_BAND_MIN) & (df['entry_price'] <= PRICE_BAND_MAX)]
            
            # Filter for losing trades (PnL < threshold, e.g., -2.5)
            df = df[df['pnl'] < PNL_THRESHOLD_LOSS]
            
            if len(df) > 0:
                # Extract date from file path and add as column
                trade_date = extract_date_from_file_path(trade_file)
                if trade_date:
                    df['date'] = trade_date
                else:
                    df['date'] = None
                    logger.warning(f"Could not extract date from {trade_file}, setting date to None")
                
                df['source_trade_file'] = str(trade_file)
                all_trades.append(df)
                logger.info(f"Loaded {len(df)} losing trades from {trade_file.name} (price band {PRICE_BAND_MIN}-{PRICE_BAND_MAX})")
        except Exception as e:
            logger.warning(f"Error reading {trade_file.name}: {e}")
            continue
    
    if not all_trades:
        logger.warning("No losing trades found in the specified price band")
        return
    
    # Combine all trades
    combined_df = pd.concat(all_trades, ignore_index=True)
    
    # Enrich with NIFTY data (prev day close, current price, sentiment)
    sentiment_enabled = analytics_sentiment_enabled(config_data)
    combined_df = enrich_with_nifty_data(combined_df, enable_sentiment=sentiment_enabled)
    
    # Enrich with high and swing low data
    combined_df = enrich_with_high_swing_low(combined_df, config_data)
    
    # Add skip_first indicator
    combined_df = add_skip_first_flags(combined_df)
    
    # Add indicator snapshot columns
    combined_df = add_indicator_snapshot(combined_df)
    
    # Reorder columns to put date and NIFTY columns in prominent positions
    if 'date' in combined_df.columns:
        cols = list(combined_df.columns)
        move_cols = []
        
        if 'date' in cols:
            cols.remove('date')
            move_cols.append('date')
        
        if sentiment_enabled:
            for col in SENTIMENT_COLUMNS:
                if col in cols:
                    cols.remove(col)
                    move_cols.append(col)
        
        for col in ALWAYS_ON_SENTIMENT_COLUMNS:
            if col in cols:
                cols.remove(col)
                move_cols.append(col)
        
        indicator_cols = []
        for col in INDICATOR_COLUMNS:
            if col in cols:
                cols.remove(col)
                indicator_cols.append(col)
        
        if 'symbol' in cols:
            insert_idx = cols.index('symbol') + 1
        elif 'symbol_html' in cols:
            insert_idx = cols.index('symbol_html') + 1
        else:
            insert_idx = 0
        
        for i, col in enumerate(move_cols):
            cols.insert(insert_idx + i, col)
        
        if 'symbol' in cols:
            insert_idx = cols.index('symbol') + 1
        elif 'symbol_html' in cols:
            insert_idx = cols.index('symbol_html') + 1
        else:
            insert_idx = 0
        
        for i, col in enumerate(indicator_cols):
            cols.insert(insert_idx + i, col)
        
        combined_df = combined_df[cols]
    
    if 'source_trade_file' in combined_df.columns:
        combined_df = combined_df.drop(columns=['source_trade_file'])
    logger.info(f"\n{'='*60}")
    logger.info(f"EXPORT SUMMARY")
    logger.info(f"{'='*60}")
    logger.info(f"Total losing trades in price band {PRICE_BAND_MIN}-{PRICE_BAND_MAX}: {len(combined_df)}")
    logger.info(f"Total PnL: {combined_df['pnl'].sum():.2f}")
    logger.info(f"{'='*60}\n")
    
    # Save to Excel with hyperlink preservation
    output_file = output_dir / 'losing_trades_70_249.xlsx'
    
    # Check if file is locked/open before attempting to write
    if output_file.exists():
        try:
            # Try to open with openpyxl to check if file is locked
            from openpyxl import load_workbook
            try:
                wb_test = load_workbook(output_file, read_only=True)
                wb_test.close()
            except PermissionError:
                raise
            except Exception:
                # File might be corrupted or not a valid Excel file, but not locked
                pass
        except PermissionError:
            logger.warning("=" * 60)
            logger.warning("WARNING: Excel file is currently open or locked!")
            logger.warning(f"File: {output_file}")
            logger.warning("Please close the file in Excel or other programs and rerun the script.")
            logger.warning("Falling back to CSV format...")
            logger.warning("=" * 60)
            output_file_csv = output_dir / 'losing_trades_70_249.csv'
            combined_df.to_csv(output_file_csv, index=False)
            logger.info(f"Saved as CSV instead: {output_file_csv}")
            return
    
    try:
        from openpyxl import load_workbook
        import re
        
        # First save to Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            combined_df.to_excel(writer, sheet_name='Losing Trades', index=False)
        
        # Now preserve hyperlinks by reading the CSV-style hyperlinks and converting to Excel hyperlinks
        wb = load_workbook(output_file)
        ws = wb['Losing Trades']
        
        # Find symbol and symbol_html column indices
        header_row = 1
        symbol_col = None
        symbol_html_col = None
        
        for col_idx, cell in enumerate(ws[header_row], 1):
            if cell.value == 'symbol':
                symbol_col = col_idx
            elif cell.value == 'symbol_html':
                symbol_html_col = col_idx
        
        # Process hyperlinks in symbol column
        if symbol_col:
            for row_idx in range(2, len(combined_df) + 2):  # Start from row 2 (after header)
                cell = ws.cell(row=row_idx, column=symbol_col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('=HYPERLINK('):
                    # Extract URL and display text from HYPERLINK formula
                    # Handle escaped quotes in CSV format: =HYPERLINK(""path"", ""text"")
                    match = re.search(r'=HYPERLINK\(""([^"]+)"",\s*""([^"]+)""\)', cell.value)
                    if not match:
                        # Try standard format: =HYPERLINK("path", "text")
                        match = re.search(r'=HYPERLINK\("([^"]+)",\s*"([^"]+)"\)', cell.value)
                    if match:
                        url = match.group(1)
                        display_text = match.group(2)
                        cell.value = display_text
                        # Create hyperlink using openpyxl's hyperlink module
                        from openpyxl.worksheet.hyperlink import Hyperlink
                        cell.hyperlink = Hyperlink(ref=cell.coordinate, target=url, tooltip=url)
                        cell.style = "Hyperlink"
        
        # Process hyperlinks in symbol_html column
        if symbol_html_col:
            for row_idx in range(2, len(combined_df) + 2):
                cell = ws.cell(row=row_idx, column=symbol_html_col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('=HYPERLINK('):
                    # Extract URL and display text from HYPERLINK formula
                    match = re.search(r'=HYPERLINK\(""([^"]+)"",\s*""([^"]+)""\)', cell.value)
                    if not match:
                        match = re.search(r'=HYPERLINK\("([^"]+)",\s*"([^"]+)"\)', cell.value)
                    if match:
                        url = match.group(1)
                        display_text = match.group(2)
                        cell.value = display_text
                        # Create hyperlink using openpyxl's hyperlink module
                        from openpyxl.worksheet.hyperlink import Hyperlink
                        cell.hyperlink = Hyperlink(ref=cell.coordinate, target=url, tooltip=url)
                        cell.style = "Hyperlink"
        
        wb.save(output_file)
        logger.info(f"Saved losing trades to {output_file}")
        logger.info(f"Total rows: {len(combined_df)}")
    except PermissionError as e:
        logger.error("=" * 60)
        logger.error("ERROR: Permission denied while saving Excel file!")
        logger.error(f"File: {output_file}")
        logger.error("The file is likely open in Excel or locked by another process.")
        logger.error("Please close the file and rerun the script, or use the CSV file that was created.")
        logger.error("=" * 60)
        import traceback
        logger.debug(traceback.format_exc())
        # Fallback to CSV if Excel fails
        output_file_csv = output_dir / 'losing_trades_70_249.csv'
        combined_df.to_csv(output_file_csv, index=False)
        logger.info(f"Saved as CSV instead: {output_file_csv}")
    except Exception as e:
        logger.error(f"Error saving to Excel: {e}")
        import traceback
        logger.error(traceback.format_exc())
        # Fallback to CSV if Excel fails
        output_file_csv = output_dir / 'losing_trades_70_249.csv'
        combined_df.to_csv(output_file_csv, index=False)
        logger.info(f"Saved as CSV instead: {output_file_csv}")

if __name__ == '__main__':
    main()

