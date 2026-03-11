#!/usr/bin/env python3
"""
Create Losing Trades Dataset
Consolidates complete trade cycle from entry to end for each option symbol
Based on format from create_winning_trades_dataset.py
"""

import pandas as pd
from pathlib import Path
import logging
import re
from datetime import datetime as dt
from typing import Optional

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def extract_hyperlink_path(value: str) -> Optional[str]:
    """Pull path from Excel-style HYPERLINK string."""
    if not isinstance(value, str):
        return None
    # Try multiple patterns
    # Pattern 1: =HYPERLINK("path", "text")
    # Pattern 2: =HYPERLINK(""path"", ""text"")
    patterns = [
        r'=HYPERLINK\("([^"]+)"',  # Standard
        r'=HYPERLINK\(""([^"]+)""',  # Escaped quotes
        r'=HYPERLINK\("([^"]+)",\s*"',  # With comma
    ]
    for pattern in patterns:
        match = re.search(pattern, value.replace('""', '"'))
        if match:
            return match.group(1)
    return None

def resolve_strategy_file(symbol_value, symbol_html_value, source_file: Optional[str]) -> Optional[Path]:
    """Best-effort resolution of strategy CSV path for a trade row."""
    # 1. Direct hyperlink on symbol column
    hyperlink_path = extract_hyperlink_path(symbol_value) if symbol_value else None
    if hyperlink_path:
        return Path(hyperlink_path)
    
    # 2. Hyperlink on symbol_html column (HTML file) -> convert to CSV
    html_path = extract_hyperlink_path(symbol_html_value) if symbol_html_value else None
    if html_path:
        csv_candidate = Path(html_path).with_suffix('.csv')
        return csv_candidate
    
    # 3. Construct from symbol text and source directory
    symbol_text = None
    if isinstance(symbol_value, str):
        if '=HYPERLINK' in symbol_value:
            text_match = re.findall(r'"([^"]+)"', symbol_value.replace('""', '"'))
            if len(text_match) >= 2:
                symbol_text = text_match[1]
        else:
            symbol_text = symbol_value.strip()
    
    if not symbol_text:
        return None
    
    strategy_filename = symbol_text
    if not strategy_filename.endswith('_strategy.csv'):
        strategy_filename = f"{strategy_filename}_strategy.csv"
    
    if source_file:
        trade_dir = Path(source_file).parent
        atm_dir = trade_dir / 'ATM'
        candidate_dirs = [atm_dir, trade_dir]
    else:
        candidate_dirs = []
    
    for directory in candidate_dirs:
        candidate_path = directory / strategy_filename
        if candidate_path.exists():
            return candidate_path
    
    # Fall back to direct filename if reachable
    candidate_path = Path(strategy_filename)
    return candidate_path if candidate_path.exists() else None

def normalize_time_str(time_value: Optional[str]) -> Optional[str]:
    """Ensure time string includes seconds (HH:MM:SS)."""
    if not time_value or not isinstance(time_value, str):
        return None
    time_value = time_value.strip()
    if not time_value:
        return None
    if len(time_value) == 5:
        return f"{time_value}:00"
    return time_value

def extract_trade_cycle(strategy_file: Path, entry_time: str, exit_time: str, symbol_name: str) -> Optional[pd.DataFrame]:
    """
    Extract complete trade cycle from strategy file between entry_time and exit_time.
    Returns DataFrame with all rows from entry to exit, plus symbol name in last column.
    """
    try:
        # Read strategy file
        df = pd.read_csv(strategy_file)
        if df.empty or 'date' not in df.columns:
            logger.warning(f"Strategy file {strategy_file.name} is empty or missing 'date' column")
            return None
        
        df['date'] = pd.to_datetime(df['date'])
        # Remove timezone from date if present
        if hasattr(df['date'].dtype, 'tz') and df['date'].dtype.tz is not None:
            df['date'] = df['date'].dt.tz_localize(None)
        else:
            # Check sample value
            sample = df['date'].dropna()
            if len(sample) > 0 and hasattr(sample.iloc[0], 'tz') and sample.iloc[0].tz is not None:
                df['date'] = df['date'].dt.tz_localize(None)
        
        df['time_only'] = df['date'].dt.time
        
        # Normalize entry and exit times
        entry_time_norm = normalize_time_str(entry_time)
        exit_time_norm = normalize_time_str(exit_time)
        
        if not entry_time_norm or not exit_time_norm:
            logger.warning(f"Could not normalize times: entry={entry_time}, exit={exit_time}")
            return None
        
        # Parse times
        try:
            entry_time_obj = dt.strptime(entry_time_norm, '%H:%M:%S').time()
            exit_time_obj = dt.strptime(exit_time_norm, '%H:%M:%S').time()
        except ValueError:
            logger.warning(f"Could not parse times: entry={entry_time_norm}, exit={exit_time_norm}")
            return None
        
        # Find entry row index
        entry_indices = df[df['time_only'] == entry_time_obj].index
        if len(entry_indices) == 0:
            # Try to find closest match within 5 minutes
            entry_seconds = entry_time_obj.hour * 3600 + entry_time_obj.minute * 60 + entry_time_obj.second
            min_diff = float('inf')
            entry_idx = None
            
            for idx, row in df.iterrows():
                row_time = row['time_only']
                row_seconds = row_time.hour * 3600 + row_time.minute * 60 + row_time.second
                diff = abs(row_seconds - entry_seconds)
                if diff < min_diff and diff <= 300:  # Within 5 minutes
                    min_diff = diff
                    entry_idx = idx
            
            if entry_idx is None:
                logger.warning(f"Could not find entry time {entry_time_norm} in {strategy_file.name}")
                return None
        else:
            entry_idx = entry_indices[0]
        
        # Find exit row index
        exit_indices = df[df['time_only'] == exit_time_obj].index
        if len(exit_indices) == 0:
            # Try to find closest match within 5 minutes
            exit_seconds = exit_time_obj.hour * 3600 + exit_time_obj.minute * 60 + exit_time_obj.second
            min_diff = float('inf')
            exit_idx = None
            
            for idx, row in df.iterrows():
                if idx < entry_idx:
                    continue
                row_time = row['time_only']
                row_seconds = row_time.hour * 3600 + row_time.minute * 60 + row_time.second
                diff = abs(row_seconds - exit_seconds)
                if diff < min_diff and diff <= 300:  # Within 5 minutes
                    min_diff = diff
                    exit_idx = idx
            
            if exit_idx is None:
                logger.warning(f"Could not find exit time {exit_time_norm} in {strategy_file.name}")
                return None
        else:
            # Get the first exit index after entry
            exit_indices_after_entry = exit_indices[exit_indices > entry_idx]
            if len(exit_indices_after_entry) > 0:
                exit_idx = exit_indices_after_entry[0]
            else:
                exit_idx = exit_indices[-1]  # Use last exit if no exit after entry
        
        # Extract trade cycle (entry to exit, inclusive)
        if exit_idx < entry_idx:
            logger.warning(f"Exit index {exit_idx} is before entry index {entry_idx} in {strategy_file.name}")
            return None
        
        trade_cycle = df.iloc[entry_idx:exit_idx+1].copy()
        
        # Add symbol name as last column
        trade_cycle['symbol'] = symbol_name
        
        return trade_cycle
        
    except Exception as e:
        logger.error(f"Error extracting trade cycle from {strategy_file.name}: {e}")
        import traceback
        logger.debug(traceback.format_exc())
        return None

def create_dataset():
    """Create the losing trades dataset."""
    # File paths - use relative paths from script location
    script_dir = Path(__file__).parent
    losing_trades_file = script_dir / "losing_trades_70_249.xlsx"
    output_file = script_dir / "losing_trades_dataset.xlsx"
    
    # Check if files exist
    if not losing_trades_file.exists():
        logger.error(f"Losing trades file not found: {losing_trades_file}")
        return
    
    # Read losing trades - use openpyxl to preserve hyperlinks
    logger.info(f"Reading losing trades from {losing_trades_file}")
    try:
        from openpyxl import load_workbook
        
        # First read with pandas to get data
        df_losing = pd.read_excel(losing_trades_file)
        logger.info(f"Loaded {len(df_losing)} losing trades")
        logger.info(f"Columns in losing trades: {list(df_losing.columns)[:10]}...")
        
        # Check for required columns
        required_cols = ['symbol', 'entry_time', 'exit_time']
        missing_cols = [col for col in required_cols if col not in df_losing.columns]
        if missing_cols:
            logger.error(f"Missing required columns: {missing_cols}")
            logger.error(f"Available columns: {list(df_losing.columns)}")
            return
        
        # Now read with openpyxl to extract hyperlinks from symbol column
        # Note: read_only mode doesn't support hyperlink attribute, so we use keep_links=True
        logger.info("Extracting hyperlinks from symbol column using openpyxl...")
        wb = load_workbook(losing_trades_file, read_only=False, keep_links=True)
        ws = wb.active
        
        # Find symbol column index
        symbol_col_idx = None
        for col_idx, cell in enumerate(ws[1], 1):  # First row is header
            if cell.value == 'symbol':
                symbol_col_idx = col_idx
                break
        
        # Extract hyperlinks and update dataframe
        hyperlink_paths = {}
        if symbol_col_idx:
            logger.info(f"Found symbol column at index {symbol_col_idx}")
            for row_idx in range(2, len(df_losing) + 2):  # Start from row 2 (skip header)
                cell = ws.cell(row=row_idx, column=symbol_col_idx)
                # Check if cell has hyperlink
                if hasattr(cell, 'hyperlink') and cell.hyperlink:
                    target = cell.hyperlink.target
                    if target:
                        # Clean up hyperlink target - remove "file:\" or "file:///" prefix
                        original_target = target
                        if target.startswith('file:\\') or target.startswith('file:///'):
                            # Remove file: protocol prefix
                            target = target.replace('file:\\', '').replace('file:///', '')
                            # Remove leading slashes but keep the drive letter
                            if target.startswith('\\'):
                                target = target.lstrip('\\')
                        
                        # Hyperlinks might be relative or absolute
                        # Try as-is first
                        target_path = Path(target)
                        
                        logger.debug(f"Row {row_idx-1}: Hyperlink target: {original_target} -> cleaned: {target}")
                        
                        # If relative, try to make it absolute
                        if not target_path.is_absolute():
                            project_root = Path(__file__).parent.parent.parent
                            # Try direct
                            abs_target = project_root / target.lstrip('/\\')
                            if abs_target.exists():
                                target_path = abs_target
                            else:
                                # Try with backtesting/data prefix
                                abs_target = project_root / 'backtesting' / 'data' / target.lstrip('/\\')
                                if abs_target.exists():
                                    target_path = abs_target
                                else:
                                    # Try as absolute from project root
                                    abs_target = project_root / target.replace('\\', '/').lstrip('/')
                                    if abs_target.exists():
                                        target_path = abs_target
                        
                        # Convert HTML to CSV if needed
                        if target_path.suffix == '.html':
                            target_path = target_path.with_suffix('.csv')
                        
                        # Store the absolute path
                        hyperlink_paths[row_idx - 2] = str(target_path)
                        logger.debug(f"Row {row_idx-1}: Extracted hyperlink: {target} -> {target_path} (exists: {target_path.exists()})")
            
            # Add hyperlink paths to dataframe
            if hyperlink_paths:
                df_losing['symbol_hyperlink'] = df_losing.index.map(hyperlink_paths)
                logger.info(f"✓ Extracted {len(hyperlink_paths)} hyperlinks from {len(df_losing)} rows")
            else:
                logger.warning("⚠ No hyperlinks found in symbol column")
        else:
            logger.warning("⚠ Symbol column not found in Excel file")
        
        wb.close()
        
    except Exception as e:
        logger.error(f"Error reading losing trades file: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return
    
    # Process each trade - use dictionary to ensure one-to-one mapping
    trade_cycles_dict = {}  # Map input index -> trade cycle DataFrame
    missing_strategy_files = []
    
    # Try to find source_trade_file column or date column to help locate strategy files
    has_source_file_col = 'source_trade_file' in df_losing.columns
    has_date_col = 'date' in df_losing.columns
    
    logger.info(f"Columns available: source_trade_file={has_source_file_col}, date={has_date_col}")
    logger.info(f"Processing {len(df_losing)} trades from input file...")
    
    for idx, row in df_losing.iterrows():
        logger.debug(f"Processing row {idx} (trade {len(trade_cycles_dict) + 1}/{len(df_losing)})")
        symbol_value = row.get('symbol')
        symbol_hyperlink = row.get('symbol_hyperlink', None)  # From openpyxl extraction
        symbol_html_value = row.get('symbol_html', None)
        source_file = row.get('source_trade_file', None) if has_source_file_col else None
        entry_time = row.get('entry_time')
        exit_time = row.get('exit_time')
        date_value = row.get('date', None) if has_date_col else None
        
        # Extract symbol name and strategy file path
        symbol_name = None
        strategy_file = None
        
        # First priority: Use hyperlink extracted from openpyxl
        if symbol_hyperlink:
            strategy_file = Path(symbol_hyperlink)
            
            # Verify the file exists
            if strategy_file.exists():
                symbol_name = strategy_file.stem.replace('_strategy', '')
                logger.info(f"Row {idx+1}/{len(df_losing)}: ✓ Found strategy file via hyperlink: {strategy_file.name}")
            else:
                # The hyperlink path should already be absolute from extraction, but double-check
                logger.warning(f"Row {idx+1}: Hyperlink path doesn't exist: {symbol_hyperlink}")
                # Extract symbol name for fallback search
                symbol_name = strategy_file.stem.replace('_strategy', '')
                strategy_file = None  # Reset to try fallback
        
        # Second priority: Extract from symbol_value if it contains HYPERLINK
        if not strategy_file or not strategy_file.exists():
            if isinstance(symbol_value, str):
                if '=HYPERLINK' in symbol_value:
                    # Try to extract the file path directly from hyperlink
                    hyperlink_path = extract_hyperlink_path(symbol_value)
                    if hyperlink_path:
                        strategy_file = Path(hyperlink_path)
                        # If it's HTML, convert to CSV
                        if hyperlink_path.endswith('.html'):
                            strategy_file = strategy_file.with_suffix('.csv')
                        
                        # Extract symbol name from the path
                        if not symbol_name:
                            if strategy_file.exists():
                                symbol_name = strategy_file.stem.replace('_strategy', '')
                            else:
                                symbol_name = strategy_file.stem.replace('_strategy', '')
                    
                    # If we don't have symbol_name yet, extract from display text
                    if not symbol_name:
                        text_matches = re.findall(r'"([^"]+)"', symbol_value.replace('""', '"'))
                        if len(text_matches) >= 2:
                            # Second match is usually the display text
                            symbol_name = text_matches[1]
                            # Remove _strategy.csv suffix if present
                            if symbol_name.endswith('_strategy.csv'):
                                symbol_name = symbol_name.replace('_strategy.csv', '')
                        elif len(text_matches) == 1 and not hyperlink_path:
                            symbol_name = text_matches[0]
                            if symbol_name.endswith('_strategy.csv'):
                                symbol_name = symbol_name.replace('_strategy.csv', '')
                else:
                    symbol_name = symbol_value.strip()
        
        if not symbol_name:
            logger.warning(f"Could not extract symbol name from row {idx} - using symbol_value as fallback")
            # Use symbol_value as fallback, or create a generic placeholder
            if isinstance(symbol_value, str) and symbol_value.strip():
                symbol_name = symbol_value.strip()
            else:
                symbol_name = f"UNKNOWN_SYMBOL_{idx}"
                logger.warning(f"  Using fallback symbol name: {symbol_name}")
        
        # If strategy file not found from hyperlink, try resolve_strategy_file
        if not strategy_file or not strategy_file.exists():
            strategy_file = resolve_strategy_file(symbol_value, symbol_html_value, source_file)
        
        # If still not found, try searching in data directories using the correct structure
        # Structure: backtesting/data/{expiry}_DYNAMIC/{day}/ATM/{symbol}_strategy.csv
        if not strategy_file or not strategy_file.exists():
            project_root = Path(__file__).parent.parent.parent
            data_dir = project_root / 'backtesting' / 'data'
            
            strategy_filename = f"{symbol_name}_strategy.csv"
            
            # If we have a date, try to narrow down the search
            day_label = None
            if date_value:
                try:
                    if isinstance(date_value, str):
                        date_obj = pd.to_datetime(date_value).date()
                    else:
                        date_obj = pd.to_datetime(date_value).date()
                    
                    # Convert date to day label format (e.g., NOV26, OCT15)
                    # Format: MONTH + DAY (remove leading zero from day if < 10)
                    day_str = str(date_obj.day)
                    if day_str.startswith('0'):
                        day_str = day_str[1:]  # Remove leading zero
                    day_label = date_obj.strftime('%b').upper() + day_str
                    
                    logger.debug(f"Row {idx}: Searching for {symbol_name} on date {date_obj} (day_label: {day_label})")
                except Exception as e:
                    logger.debug(f"Row {idx}: Could not parse date {date_value}: {e}")
                    day_label = None
            
            if data_dir.exists():
                # Search in all expiry directories
                for expiry_dir in sorted(data_dir.glob('*_DYNAMIC')):
                    if day_label:
                        # Try specific day directory first
                        day_dir = expiry_dir / day_label
                        if day_dir.exists() and day_dir.is_dir():
                            # Check ATM directory first (most common location)
                            atm_dir = day_dir / 'ATM'
                            if atm_dir.exists():
                                candidate = atm_dir / strategy_filename
                                if candidate.exists():
                                    strategy_file = candidate
                                    logger.info(f"Row {idx}: ✓ Found strategy file using date: {strategy_file}")
                                    break
                            
                            # Also check day directory directly
                            if not strategy_file or not strategy_file.exists():
                                candidate = day_dir / strategy_filename
                                if candidate.exists():
                                    strategy_file = candidate
                                    logger.info(f"Row {idx}: ✓ Found strategy file in day dir: {strategy_file}")
                                    break
                        
                        if strategy_file and strategy_file.exists():
                            break
                    
                    # If not found with date, search all day directories in this expiry
                    if not strategy_file or not strategy_file.exists():
                        for day_dir in sorted(expiry_dir.glob('*')):
                            if day_dir.is_dir() and not day_dir.name.startswith('.'):
                                # Check ATM directory first
                                atm_dir = day_dir / 'ATM'
                                if atm_dir.exists():
                                    candidate = atm_dir / strategy_filename
                                    if candidate.exists():
                                        strategy_file = candidate
                                        logger.info(f"Row {idx}: ✓ Found strategy file: {strategy_file}")
                                        break
                                
                                # Also check day directory
                                if not strategy_file or not strategy_file.exists():
                                    candidate = day_dir / strategy_filename
                                    if candidate.exists():
                                        strategy_file = candidate
                                        logger.info(f"Row {idx}: ✓ Found strategy file: {strategy_file}")
                                        break
                            
                            if strategy_file and strategy_file.exists():
                                break
                    
                    if strategy_file and strategy_file.exists():
                        break
        
        # Extract trade cycle
        logger.info(f"Processing trade {idx+1}/{len(df_losing)}: {symbol_name}")
        
        if not strategy_file or not strategy_file.exists():
            logger.warning(f"Strategy file not found for {symbol_name} (row {idx}) - creating placeholder")
            missing_strategy_files.append(symbol_name)
            # Create a placeholder row with basic info from input
            placeholder_data = {
                'symbol': symbol_name,
                'entry_time': entry_time,
                'exit_time': exit_time,
            }
            # Add date if available
            if date_value:
                placeholder_data['date'] = date_value
            # Create a minimal DataFrame with symbol as last column
            placeholder_df = pd.DataFrame([placeholder_data])
            trade_cycles_dict[idx] = placeholder_df
            logger.info(f"  Created placeholder row for {symbol_name}")
            continue
        
        trade_cycle = extract_trade_cycle(strategy_file, entry_time, exit_time, symbol_name)
        
        if trade_cycle is not None and len(trade_cycle) > 0:
            trade_cycles_dict[idx] = trade_cycle
            logger.info(f"  Extracted {len(trade_cycle)} rows for {symbol_name}")
        else:
            logger.warning(f"  Could not extract trade cycle for {symbol_name} - creating placeholder")
            # Create placeholder even when extraction fails
            placeholder_data = {
                'symbol': symbol_name,
                'entry_time': entry_time,
                'exit_time': exit_time,
            }
            if date_value:
                placeholder_data['date'] = date_value
            placeholder_df = pd.DataFrame([placeholder_data])
            trade_cycles_dict[idx] = placeholder_df
            missing_strategy_files.append(symbol_name)
    
    # CRITICAL: Ensure we have exactly one trade cycle for each input row
    all_input_indices = list(df_losing.index)
    logger.info(f"Total input rows: {len(all_input_indices)}")
    logger.info(f"Trade cycles in dictionary: {len(trade_cycles_dict)}")
    
    missing_indices = [idx for idx in all_input_indices if idx not in trade_cycles_dict]
    
    if missing_indices:
        logger.error(f"⚠️  Found {len(missing_indices)} missing trades - adding placeholders now")
        logger.error(f"   Missing indices: {missing_indices[:10]}..." if len(missing_indices) > 10 else f"   Missing indices: {missing_indices}")
        for missing_idx in missing_indices:
            row = df_losing.loc[missing_idx]
            symbol_val = row.get('symbol', f'PLACEHOLDER_{missing_idx}')
            if pd.isna(symbol_val) or not str(symbol_val).strip():
                symbol_val = f'PLACEHOLDER_{missing_idx}'
            placeholder_data = {
                'symbol': str(symbol_val).strip(),
                'entry_time': str(row.get('entry_time', '')),
                'exit_time': str(row.get('exit_time', '')),
            }
            if has_date_col and 'date' in row:
                placeholder_data['date'] = row.get('date', '')
            placeholder_df = pd.DataFrame([placeholder_data])
            trade_cycles_dict[missing_idx] = placeholder_df
            logger.error(f"  ✓ Added placeholder for index {missing_idx}: {symbol_val}")
    
    # Final check before conversion
    if len(trade_cycles_dict) != len(all_input_indices):
        logger.error(f"⚠️  CRITICAL: Dictionary has {len(trade_cycles_dict)} entries but need {len(all_input_indices)}")
        logger.error("   This should not happen - adding remaining placeholders...")
        for idx in all_input_indices:
            if idx not in trade_cycles_dict:
                row = df_losing.loc[idx]
                symbol_val = row.get('symbol', f'EMERGENCY_PLACEHOLDER_{idx}')
                placeholder_data = {
                    'symbol': str(symbol_val).strip() if pd.notna(symbol_val) else f'EMERGENCY_PLACEHOLDER_{idx}',
                    'entry_time': str(row.get('entry_time', '')),
                    'exit_time': str(row.get('exit_time', '')),
                }
                if has_date_col and 'date' in row:
                    placeholder_data['date'] = row.get('date', '')
                placeholder_df = pd.DataFrame([placeholder_data])
                trade_cycles_dict[idx] = placeholder_df
                logger.error(f"   ✓ Emergency placeholder for index {idx}")
    
    # Convert dictionary to list in input order - MUST have same length
    all_trade_cycles = []
    for idx in all_input_indices:
        if idx in trade_cycles_dict:
            all_trade_cycles.append(trade_cycles_dict[idx])
        else:
            logger.error(f"   ⚠️  Index {idx} still missing after all checks!")
            # Last resort placeholder
            row = df_losing.loc[idx]
            symbol_val = row.get('symbol', f'LAST_RESORT_{idx}')
            placeholder_data = {
                'symbol': str(symbol_val).strip() if pd.notna(symbol_val) else f'LAST_RESORT_{idx}',
                'entry_time': str(row.get('entry_time', '')),
                'exit_time': str(row.get('exit_time', '')),
            }
            if has_date_col and 'date' in row:
                placeholder_data['date'] = row.get('date', '')
            placeholder_df = pd.DataFrame([placeholder_data])
            all_trade_cycles.append(placeholder_df)
    
    if not all_trade_cycles:
        logger.error("No trade cycles processed. Cannot create dataset.")
        logger.error(f"Processed {len(df_losing)} trades, but created 0 trade cycles (including placeholders).")
        return
    
    # Final verification
    if len(all_trade_cycles) != len(df_losing):
        logger.error(f"⚠️  CRITICAL ERROR: Trade count still mismatched!")
        logger.error(f"   Expected: {len(df_losing)} trades")
        logger.error(f"   Got: {len(all_trade_cycles)} trades")
        return
    else:
        logger.info(f"✅ Successfully processed all {len(all_trade_cycles)} trade cycles (matches input: {len(df_losing)})")
    if missing_strategy_files:
        logger.warning(f"Note: {len(missing_strategy_files)} trades created as placeholders due to missing strategy files or extraction failures")
    
    # Get all unique columns from all trade cycles
    all_columns = set()
    for trade_cycle in all_trade_cycles:
        all_columns.update(trade_cycle.columns)
    
    # Ensure symbol is last
    all_columns = sorted([col for col in all_columns if col != 'symbol']) + ['symbol']
    
    # Align all trade cycles to have the same columns
    aligned_cycles = []
    for trade_cycle in all_trade_cycles:
        # Add missing columns with empty values
        for col in all_columns:
            if col not in trade_cycle.columns:
                trade_cycle[col] = ''
        # Reorder columns
        trade_cycle = trade_cycle[all_columns]
        aligned_cycles.append(trade_cycle)
    
    # Final check: We MUST have exactly len(df_losing) trade cycles before combining
    if len(aligned_cycles) != len(df_losing):
        logger.error(f"⚠️  CRITICAL: Before combining, we have {len(aligned_cycles)} cycles but need {len(df_losing)}")
        logger.error("This should not happen - aborting to prevent data loss!")
        return
    
    # Combine all trade cycles with blank rows between trades
    logger.info("Combining trade cycles...")
    logger.info(f"Combining {len(aligned_cycles)} trade cycles (expected {len(df_losing)})")
    combined_rows = []
    
    for i, trade_cycle in enumerate(aligned_cycles):
        # Add trade cycle rows
        combined_rows.append(trade_cycle)
        
        # Add blank row between trades (except after last trade)
        if i < len(aligned_cycles) - 1:
            # Create a blank row with same columns
            blank_row = pd.DataFrame([{col: '' for col in all_columns}])
            combined_rows.append(blank_row)
    
    # Concatenate all rows
    final_df = pd.concat(combined_rows, ignore_index=True)
    
    # Final verification: Count unique symbols (excluding blank rows)
    non_blank_final = final_df[final_df['symbol'].notna() & (final_df['symbol'].astype(str).str.strip() != '')]
    unique_symbols = non_blank_final['symbol'].nunique()
    logger.info(f"Final dataset: {len(final_df)} total rows, {unique_symbols} unique symbols (expected {len(df_losing)})")
    
    if unique_symbols != len(df_losing):
        logger.error(f"⚠️  CRITICAL WARNING: Unique symbol count ({unique_symbols}) doesn't match input ({len(df_losing)})")
        logger.error(f"   This means {len(df_losing) - unique_symbols} trades are missing from the output!")
        
        # Get input symbols
        input_symbols = set(str(s) for s in df_losing['symbol'].unique() if pd.notna(s))
        output_symbols = set(str(s) for s in non_blank_final['symbol'].unique() if pd.notna(s))
        missing_symbols = input_symbols - output_symbols
        
        if missing_symbols:
            logger.error(f"   Missing symbols ({len(missing_symbols)}): {sorted(list(missing_symbols))[:10]}")
            if len(missing_symbols) > 10:
                logger.error(f"   ... and {len(missing_symbols) - 10} more")
        
        # Check if we can fix this by adding missing symbols as placeholders
        logger.error("   Attempting to add missing symbols as placeholders...")
        for missing_sym in missing_symbols:
            # Find the corresponding input row
            matching_rows = df_losing[df_losing['symbol'].astype(str) == str(missing_sym)]
            if len(matching_rows) > 0:
                row = matching_rows.iloc[0]
                placeholder_data = {col: '' for col in all_columns}
                placeholder_data['symbol'] = str(missing_sym)
                placeholder_data['entry_time'] = str(row.get('entry_time', ''))
                placeholder_data['exit_time'] = str(row.get('exit_time', ''))
                if 'date' in all_columns and has_date_col:
                    placeholder_data['date'] = row.get('date', '')
                placeholder_df = pd.DataFrame([placeholder_data])
                # Add before the last blank row (if any) or at the end
                final_df = pd.concat([final_df, placeholder_df], ignore_index=True)
                logger.error(f"   ✓ Added placeholder for {missing_sym}")
        
        # Re-verify
        non_blank_final = final_df[final_df['symbol'].notna() & (final_df['symbol'].astype(str).str.strip() != '')]
        unique_symbols = non_blank_final['symbol'].nunique()
        logger.info(f"   After fix: {unique_symbols} unique symbols (expected {len(df_losing)})")
    
    logger.info(f"Final dataset shape: {final_df.shape}")
    logger.info(f"Columns: {list(final_df.columns)}")
    
    # Ensure symbol is the last column
    if 'symbol' in final_df.columns:
        cols = [col for col in final_df.columns if col != 'symbol'] + ['symbol']
        final_df = final_df[cols]
    
    # Remove timezones from datetime columns (Excel doesn't support timezone-aware datetimes)
    logger.info("Removing timezones from datetime columns...")
    datetime_cols = []
    for col in final_df.columns:
        if pd.api.types.is_datetime64_any_dtype(final_df[col]):
            datetime_cols.append(col)
    
    if datetime_cols:
        logger.info(f"Found {len(datetime_cols)} datetime columns: {datetime_cols}")
        for col in datetime_cols:
            try:
                # Method 1: Try tz_localize(None) if timezone-aware
                try:
                    if hasattr(final_df[col].dtype, 'tz') and final_df[col].dtype.tz is not None:
                        final_df[col] = final_df[col].dt.tz_localize(None)
                        logger.info(f"  Removed timezone from {col} (method 1)")
                    else:
                        # Method 2: Check if values have timezone
                        sample = final_df[col].dropna()
                        if len(sample) > 0:
                            test_val = sample.iloc[0]
                            if hasattr(test_val, 'tz') and test_val.tz is not None:
                                # Use tz_localize(None) - this removes timezone from timezone-aware
                                final_df[col] = final_df[col].dt.tz_localize(None)
                                logger.info(f"  Removed timezone from {col} (method 2)")
                            else:
                                # Method 3: Convert to string and back (safest)
                                mask = final_df[col].notna()
                                if mask.any():
                                    # Format as string without timezone, then parse back
                                    final_df.loc[mask, col] = pd.to_datetime(
                                        final_df.loc[mask, col].dt.strftime('%Y-%m-%d %H:%M:%S'),
                                        format='%Y-%m-%d %H:%M:%S',
                                        errors='coerce'
                                    )
                                logger.info(f"  Removed timezone from {col} (method 3)")
                except Exception as e1:
                    logger.warning(f"  Method 1-3 failed for {col}: {e1}")
                    # Method 4: Convert to string, then back to datetime
                    try:
                        final_df[col] = pd.to_datetime(final_df[col].astype(str), errors='coerce')
                        logger.info(f"  Removed timezone from {col} (method 4)")
                    except Exception as e2:
                        logger.error(f"  All methods failed for {col}: {e2}")
                        # Last resort: keep as string
                        final_df[col] = final_df[col].astype(str)
                        logger.warning(f"  Converted {col} to string to avoid error")
            except Exception as e:
                logger.error(f"Error processing {col}: {e}")
                # Convert to string as last resort
                final_df[col] = final_df[col].astype(str)
    
    # Final verification: check for any remaining timezone-aware values
    logger.info("Final verification of datetime columns...")
    for col in final_df.columns:
        if pd.api.types.is_datetime64_any_dtype(final_df[col]):
            sample = final_df[col].dropna()
            if len(sample) > 0:
                test_val = sample.iloc[0]
                if hasattr(test_val, 'tz') and test_val.tz is not None:
                    logger.warning(f"  {col} still has timezone - converting to string")
                    final_df[col] = final_df[col].astype(str)
    
    # ULTIMATE FIX: Convert ALL datetime columns to string format before saving
    # This ensures Excel compatibility regardless of timezone issues
    logger.info("Converting all datetime columns to string format for Excel compatibility...")
    datetime_cols_to_convert = []
    for col in final_df.columns:
        if pd.api.types.is_datetime64_any_dtype(final_df[col]):
            datetime_cols_to_convert.append(col)
    
    if datetime_cols_to_convert:
        logger.info(f"Converting {len(datetime_cols_to_convert)} datetime columns to string: {datetime_cols_to_convert}")
        for col in datetime_cols_to_convert:
            try:
                # Convert entire column to string format
                # Handle both timezone-aware and naive datetimes
                def convert_datetime(val):
                    if pd.isna(val):
                        return ''
                    try:
                        # If it's already a datetime, format it
                        if isinstance(val, pd.Timestamp):
                            # Remove timezone if present, then format
                            if val.tz is not None:
                                val = val.tz_localize(None)
                            return val.strftime('%Y-%m-%d %H:%M:%S')
                        else:
                            return str(val)
                    except:
                        return str(val) if pd.notna(val) else ''
                
                final_df[col] = final_df[col].apply(convert_datetime)
                logger.info(f"  ✓ Converted {col} to string format")
            except Exception as e:
                logger.error(f"  ✗ Error converting {col}: {e}")
                # Last resort: simple string conversion
                final_df[col] = final_df[col].astype(str)
                logger.warning(f"  Converted {col} using simple string conversion")
    
    # Save to Excel
    logger.info(f"Saving dataset to {output_file}")
    try:
        # Check if file is locked
        if output_file.exists():
            try:
                from openpyxl import load_workbook
                wb_test = load_workbook(output_file, read_only=True)
                wb_test.close()
            except PermissionError:
                logger.warning(f"Output file is locked. Please close {output_file} and rerun.")
                return
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            final_df.to_excel(writer, sheet_name='Losing Trades Dataset', index=False)
        
        success_msg = f"✅ SUCCESS: Dataset saved to {output_file}"
        logger.info(success_msg)
        print("\n" + "=" * 70)
        print(success_msg)
        print("=" * 70)
        print(f"   Total rows: {len(final_df)}")
        print(f"   Total trades: {len(all_trade_cycles)}")
        print(f"   Columns: {len(final_df.columns)}")
        print(f"   Last column: {final_df.columns[-1]}")
        if missing_strategy_files:
            print(f"   ⚠ Missing strategy files: {len(missing_strategy_files)} trades")
        print("=" * 70 + "\n")
        
        logger.info(f"   Total rows: {len(final_df)}")
        logger.info(f"   Total trades: {len(all_trade_cycles)}")
        if missing_strategy_files:
            logger.warning(f"   Missing strategy files: {len(missing_strategy_files)} trades")
        
    except PermissionError as e:
        logger.error(f"Permission denied saving to {output_file}. File may be open in Excel.")
        logger.error("Please close the file and rerun the script.")
    except Exception as e:
        logger.error(f"Error saving dataset: {e}")
        import traceback
        logger.error(traceback.format_exc())

if __name__ == '__main__':
    import sys
    
    # Also log to file (append mode to see latest run)
    log_file = Path(__file__).parent / 'create_losing_dataset.log'
    file_handler = logging.FileHandler(log_file, mode='a', encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logger.addHandler(file_handler)
    
    # Also print to console
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logger.addHandler(console_handler)
    
    try:
        print("=" * 70)
        print("STARTING LOSING TRADES DATASET CREATION")
        print("=" * 70)
        logger.info("=" * 70)
        logger.info("STARTING LOSING TRADES DATASET CREATION")
        logger.info("=" * 70)
        create_dataset()
        print("=" * 70)
        print("LOSING TRADES DATASET CREATION COMPLETE")
        print("=" * 70)
        logger.info("=" * 70)
        logger.info("LOSING TRADES DATASET CREATION COMPLETE")
        logger.info("=" * 70)
    except Exception as e:
        error_msg = f"Fatal error: {e}"
        logger.error(error_msg, exc_info=True)
        print(f"\nFATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

